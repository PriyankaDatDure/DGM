"""Extract French labels from Dictionnaire sheet into JSON for i18n."""
import json
import re
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\Admin\Downloads\Template_conceptuel_formulaire_DGM_previsions_meteo_RCA_FR_regions 1.xlsx")
OUT = Path(__file__).parent / "dictionnaire-fr.json"

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)

# Find dictionary sheet (French name may vary)
sheet_name = next((s for s in wb.sheetnames if "ictionnaire" in s.lower() or s.lower() == "dictionary"), None)
if not sheet_name:
    raise SystemExit("Dictionnaire sheet not found")
ws = wb[sheet_name]
rows = list(ws.iter_rows(values_only=True))
print(f"Using sheet: {sheet_name}, rows: {len(rows)}")

# Print header row for inspection
for i, row in enumerate(rows[:8]):
    print(i, row)

entries = []
for row in rows[1:]:
    if not row or all(c is None or str(c).strip() == "" for c in row):
        continue
    cells = [str(c).strip() if c is not None else "" for c in row]
    entries.append(cells)

out = {"sheet": sheet_name, "headers": [str(c) if c else "" for c in rows[0]], "rows": entries}
OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {len(entries)} rows to {OUT}")
